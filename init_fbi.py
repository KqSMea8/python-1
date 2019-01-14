VERSION = 8.5 # 2019-1-14

# 公共变量
DBINFO = 'data/init_fbi.xlsm' # Shian
# DBINFO = r'D:\工作资料\配置表\数据库信息.xlsx' # Dixuan
# DBINFO = r'C:\Users\admin\Documents\20180720\配置表\数据库信息.xlsx'# Yawei
FDB_CONFIG = '1PUJBtHL7EUWGrPuBelCKZbUI6zZxJ-wOC0m74rjsR5o' # fdb_config_v8
FDB_CONFIG_TAB = 'config'
FDB_INPUT = '1uRjeyc5KqjM1x7igLpse14kPojw057t6TP0HwaealAA' # fdb_input_v8

# 环境变量
import os
os.environ['http_proxy'] = 'http://bj-rd-proxy.byted.org:3128'
os.environ['https_proxy'] = 'http://bj-rd-proxy.byted.org:3128'

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import time
import os
import csv
import json
import chardet
import requests

import pymysql
pymysql.install_as_MySQLdb()
from sqlalchemy import create_engine
import sqlite3

# Excel相关函数------------------------------------------------------------------------
def getExcel(filename, sheet_name, header=0, index_col=0, skiprows=0, skipfooter=0,
    returnDf=True, data_only=True):
    '''
    If returnDf=False, return a list in the format of [[r1c1, r1c2, ...], ...]
    If data_only=True (only works when returnDf=False), return formula instead of value
    '''
    if returnDf:
        xlsx = pd.ExcelFile(filename)
        df = pd.read_excel(io=xlsx,
            sheet_name=sheet_name,
            header=header,
            index_col=index_col,
            skiprows=skiprows,
            skipfooter=skipfooter)
        xlsx.close()
        return df
    else:
        sht = openpyxl.load_workbook(filename, read_only=True,
            keep_vba=False, data_only=data_only, guess_types=False, keep_links=True)
        tab = sht[sheet_name]
        '''
        from openpyxl.formula import Tokenizer
        tok = Tokenizer("""=IF($A$1,"then True",MAX(DEFAULT_VAL,'Sheet 2'!B1))""")
        print("\n".join("%12s%11s%9s" % (t.value, t.type, t.subtype) for t in tok.items))

        sht['AQ6'].value, sht.cell(6, 43).value
        usedrange = sht.calculate_dimension(force=True)
        '''
        tab_data = [[cell.value for cell in row] for row in tab]
        sht.close()
        return tab_data

def getExcelRange(filename, sheet_name, address=None,
    returnDf=True, data_only=True):
    '''
    If returnDf=False, return a list in the format of [[r1c1, r1c2, ...], ...]
    If returnDf=True, the header is always the first row and index is always None
    If data_only=True (only works when returnDf=False), return formula instead of value
    address is in A1:B2 format, can be written as A:B for columns, or 1:2 for rows
    '''
    if returnDf and address is None:
        xlsx = pd.ExcelFile(filename)
        df = pd.read_excel(io=xlsx,
            sheet_name=sheet_name,
            header=0,
            index_col=None,
            skiprows=0,
            skipfooter=0)
        xlsx.close()
        return df
    else:
        sht = openpyxl.load_workbook(filename, read_only=True,
            keep_vba=False, data_only=data_only, guess_types=False, keep_links=True)
        tab = sht[sheet_name]
        '''
        from openpyxl.formula import Tokenizer
        tok = Tokenizer("""=IF($A$1,"then True",MAX(DEFAULT_VAL,'Sheet 2'!B1))""")
        print("\n".join("%12s%11s%9s" % (t.value, t.type, t.subtype) for t in tok.items))

        sht['AQ6'].value, sht.cell(6, 43).value
        usedrange = sht.calculate_dimension(force=True)
        '''
        if address is None:
            tab_data = [[cell.value for cell in row] for row in tab]
        else:
            address_rc = excelRangeToRC(address, asString=False)
            tab_data = []
            for rowId, row in enumerate(tab.iter_rows()):
                if (np.isnan(address_rc[0][0]) or \
                    rowId + 1 >= address_rc[0][0]) and \
                    (np.isnan(address_rc[1][0]) or \
                    rowId + 1 <= address_rc[1][0]):
                    # row is in range
                    row_list = []
                    for colId, cell in enumerate(row):
                        if (np.isnan(address_rc[0][1]) or \
                            colId + 1 >= address_rc[0][1]) and \
                            (np.isnan(address_rc[1][1]) or \
                            colId + 1 <= address_rc[1][1]):
                            # col is in range
                            row_list.append(cell.value)
                tab_data.append(row_list)
        if returnDf:
            tab_data = pd.DataFrame(data=tab_data[1:], columns=tab_data[0])
            tab_data.dropna(axis=0, how='all', inplace=True)
        sht.close()
        return tab_data

def getExcelTabSize(excel_file, tab_name):
    '''
    return the size of the sheet (tab) as (rows, columns)
    '''
    sht = openpyxl.load_workbook(excel_file, read_only=True,
        keep_vba=False, data_only=False, guess_types=False, keep_links=True)
    tab = sht[tab_name]
    dimension = tab.calculate_dimension() # returns an address as A1:AM559
    tab_size = excelRangeToRC(dimension, asString=False)[1]
    sht.close()
    return tab_size

def getExcelSheets(excelFile):
    '''
    return a list of sheet names in excelFile
    '''
    sht = openpyxl.load_workbook(excelFile, read_only=True,
        keep_vba=False, data_only=False, guess_types=False, keep_links=True)
    tab_names = sht.sheetnames
    sht.close()
    return tab_names

def isExcelTabExist(filename, sheet_name):
    return sheet_name in getExcelSheets(filename)

def writeExcelRange(df, filename, sheet_name='Sheet1', address='A1',
    header=True, index=True):
    '''
    Existing Excel filename and / or sheet_name will be preserved
    '''
    if not isFileExist(filename):
        sht = openpyxl.Workbook()
        sht.save(filename)
        print(filename, 'created')
    else:
        sht = openpyxl.load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = sht
    if sheet_name not in writer.book.sheetnames:
        writer.book.create_sheet(sheet_name)
        print(sheet_name, 'added')
    # idx = writer.book.sheetnames.index(sheet_name)
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    adresss = address.split(':')[0] # end range is omitted
    address = excelRangeToRC(address, asString=False)
    # can only use once
    # pd.io.formats.excel.header_style = {'font': {'bold': True},
    #      'borders': {'top': 'thin', 'right': 'thin', 'bottom': 'thin', 'left': 'thin'},
    #      'alignment': {'horizontal': 'center', 'vertical': 'top'}}
    # pd.io.formats.excel.header_style = None
    df.to_excel(excel_writer=writer,
        sheet_name=sheet_name,
        header=header,
        index=index,
        startrow=address[0] - 1,
        startcol=address[1] - 1)
    writer.save()
    writer.close()
    sht.close()
    print(filename, sheet_name, df.shape)

def writeExcel(df, filename, sheet_name='Sheet1', header=True, index=True,
    freeze_panes=(1, 1), create=False, resize=True):
    '''
    if create=True, always create a new Excel file and overwrite any existing file
    else append to an exsiting file or overwrite the same sheet_name
    if resize and sheet_name exists, oversize rows and columns will be deleted
    '''
    if (not create) and isFileExist(filename):
        if resize:
            sht = openpyxl.load_workbook(filename)
            if sheet_name in sht.sheetnames:
                print(sheet_name, 'exists')
                tab = sht[sheet_name]
                tab_size = getExcelTabSize(filename, sheet_name)
                header_size = len(df.columns.names) if header else 0
                index_size = len(df.index.names) if index else 0
                if tab_size[0] > df.shape[0] + header_size:
                    # oversize rows
                    tab.delete_rows(df.shape[0] + header_size + 1, tab_size[0])
                    print('rows', df.shape[0] + header_size + 1, 'to', tab_size[0], 'deleted')
                if tab_size[1] > df.shape[1] + index_size:
                    # oversize columns
                    tab.delete_cols(df.shape[1] + index_size + 1, tab_size[1])
                    print('columns', df.shape[1] + index_size + 1, 'to', tab_size[1], 'deleted')
            sht.close()
        writeExcelRange(df, filename, sheet_name=sheet_name, address='A1',
            header=header, index=index)
    else:
        writer = pd.ExcelWriter(filename)
        df.to_excel(excel_writer=writer,
            sheet_name=sheet_name,
            header=header,
            index=index,
            freeze_panes=freeze_panes)
        writer.save()
        writer.close()
        print(filename, sheet_name, df.shape)

def writeExcels(dfs, filename, header=True, index=True,
    freeze_panes=(1, 1), create=False, resize=True):
    '''
    write a list of dfs into one Excel file
    dfs: {'sheet name': df}
    if dfs is a list like [df1, df2, ...],
    sheet names will be automatically populated as Sheet1, Sheet2, ...
    set create to True to write large amount of data
    if resize and sheet_name exists, oversize rows and columns will be deleted
    '''
    if isinstance(dfs, list):
        dfdict = {}
        for i, df in enuemerate(dfs):
            dfdict['Sheet'+str(i+1)] = df
    else:
        dfdict = dfs
    if not create:
        for sheet_name, df in dfdict.items():
            writeExcel(df, filename, sheet_name=sheet_name,
                header=header, index=index, freeze_panes=freeze_panes,
                create=False, resize=resize)
    else:
        writer = pd.ExcelWriter(filename)
        for sheet_name, df in dfdict.items():
            df.to_excel(excel_writer=writer,
                sheet_name=sheet_name,
                header=header,
                index=index,
                freeze_panes=freeze_panes)
            print(filename, sheet_name, df.shape)
        writer.save()
        writer.close()

def getTxtEncoding(file, read_size=1000):
    '''
    二进制方式读取，获取字节数据，检测类型
    read_size为用于判断编码格式的字符个数，-1表示全量读取
    return a dictionary as:
    {'encoding': 'utf-8', 'confidence': 0.99, 'language': ''}
    '''
    with open(file, 'rb') as f:
        return chardet.detect(f.read(read_size))

def getCsv(filename, sep=',', quotechar='"', encoding='utf8', header=0, index_col=None,
    skiprows=0, skipfooter=0):
    '''
    default encoding changed from gbk to utf8
    '''
    f = open(filename, 'r', encoding=encoding)
    df = pd.read_csv(filepath_or_buffer=f,
        encoding=encoding,
        sep=sep,
        quotechar=quotechar,
        quoting=0, # QUOTE_MINIMAL (0), QUOTE_ALL (1), QUOTE_NONNUMERIC (2) or QUOTE_NONE (3)
        doublequote=True,
        header=header,
        index_col=index_col,
        skiprows=skiprows,
        skipfooter=skipfooter)
    f.close()
    return df

def writeCsv(df, filename, mode='a', encoding='utf8', sep=',', header=True, index=True):
    '''
    mode: 'w' for write, 'a' for append
    default encoding changed from gbk to utf8
    '''
    df.to_csv(filename,
         mode=mode,
         encoding=encoding, # default utf-8, gbk for Chinese in Excel
         sep=sep,
         header=header,
         index=index)
    print(filename, df.shape)

def concatCsvFiles(csvFiles, encoding='utf8'):
    '''
    concat and return a DataFrame
    default encoding changed from gbk to utf8
    '''
    df = None
    for file in csvFiles:
        df1 = getCsv(file, encoding=encoding)
        print(df1.shape, file)
        df = df1 if df is None else pd.concat([df, df1], sort=False)
    return df

def mergeCsvFiles(csvFiles, output='combined.csv', in_encoding='utf8', out_encoding='utf8',
    combine_header=True, skiprows=0, skipfooter=0):
    '''
    merge and write into an output csv file
    csvFiles can be a list or a single file (to convert encoding)
    use skiprows and skipfooter for cleansing (be careful if csv file is too large)
    default encoding changed from gbk to utf8
    '''
    fw = open(output, 'w', encoding=out_encoding)
    writer = csv.writer(fw)
    if not isinstance(csvFiles, list):
        csvFiles = [csvFiles]
    for idx, file in enumerate(csvFiles):
        with open(file, 'r', encoding=in_encoding) as f:
            rowid = 0
            reader = csv.reader(f)
            if skiprows>0 or skipfooter>0:
                filtered = [row for row in reader]
                if skiprows>0:
                    filtered = filtered[skiprows:]
                if skipfooter>0:
                    filtered = filtered[:-skipfooter]
                if idx!=0 and combine_header:
                    filtered = filtered[1:]
                for row in filtered:
                    writer.writerow(row)
            else:
                if idx!=0 and combine_header:
                    next(reader) # skip duplicate header
                for row in reader:
                    writer.writerow(row)
            print(file)
    fw.close()
    return True

def convertTxtEncoding(csvFiles, in_encoding='gbk', out_encoding='utf8'):
    '''
    convert each csv file from in_encoding to out_encoding
    new file will be added a post-fix of the encoding to the name
    '''
    if not isinstance(csvFiles, list):
        csvFiles = [csvFiles]
    for file in csvFiles:
        output = os.path.splitext(file)[0] + '_' + out_encoding + os.path.splitext(file)[1]
        print(file, output)
        with open(file, 'r', encoding=in_encoding) as fr:
            reader = csv.reader(fr)
            with open(output, 'w', encoding=out_encoding) as fw:
                writer = csv.writer(fw)
                for row in reader:
                    writer.writerow(row)
    return True

def xldate_to_datetime(xldate, datemode=0):
    '''
    datemode: 0 for 1900-based, 1 for 1904-based
    xldate_to_datetime(43101) == datetime_to_xldate(pd.to_datetime('20180101'))
    '''
    if isinstance(xldate, list):
        return [xldate_to_datetime(d0, datemode) for d0 in xldate]
    elif isinstance(xldate, pd.Series):
        return xldate.map(lambda d0: xldate_to_datetime(d0, datemode))
    else:
        return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=xldate + 1462 * datemode)
def datetime_to_xldate(dtt, datemode=0):
    '''
    datemode: 0 for 1900-based, 1 for 1904-based
    xldate_to_datetime(43101) == datetime_to_xldate(pd.to_datetime('20180101'))
    '''
    if isinstance(dtt, list):
        return [datetime_to_xldate(d0, datemode) for d0 in dtt]
    elif isinstance(dtt, pd.Series):
        return dtt.map(lambda d0: datetime_to_xldate(d0, datemode))
    else:
        return (dtt - datetime.datetime(1899, 12, 30)).days - 1462 * datemode

def eomonth(d, months=0):
    '''
    return end of month as datetime
    d可以为一个日期，一个含有日期的列表或者Pandas Series
    '''
    if isinstance(d, list):
        return [eomonth(d0, months) for d0 in d]
    elif isinstance(d, pd.Series):
        return d.map(lambda d0: eomonth(d0, months))
    else:
        y, m = divmod(d.month + months + 1, 12)
    #     y,m=int(y),int(m)
        if m == 0:
            y -= 1
            m = 12
        return datetime.datetime(d.year + y, m, 1) - datetime.timedelta(days=1)

def bomonth(d, months=0):
    '''
    return begin of month as datetime
    d可以为一个日期，一个含有日期的列表或者Pandas Series
    '''
    if isinstance(d, list):
        return [bomonth(d0, months) for d0 in d]
    elif isinstance(d, pd.Series):
        return d.map(lambda d0: bomonth(d0, months))
    else:
        e = eomonth(d, months)
        return datetime.datetime(e.year, e.month, 1)

def edate(d, months=0):
    '''
    return the same date in n months as datetime
    d可以为一个日期，一个含有日期的列表或者Pandas Series
    '''
    if isinstance(d, list):
        return [edate(d0, months) for d0 in d]
    elif isinstance(d, pd.Series):
        return d.map(lambda d0: edate(d0, months))
    else:
        e = eomonth(d, months)
        return datetime.datetime(e.year, e.month, min(d.day, e.day))

def months(d1, d2):
    '''
    return months between two dates, d2-d1
    '''
    if isinstance(d1, pd.Series):
        d1 = d1.tolist()
    if isinstance(d2, pd.Series):
        d2 = d2.tolist()
    if isinstance(d1, list) and not isinstance(d2, list):
        d2 = [d2] * len(d1)
    if isinstance(d2, list) and not isinstance(d1, list):
        d1 = [d1] * len(d2)
    if isinstance(d1, list):
        return [months(d3, d4) for (d3, d4) in zip(d1, d2)]
    else:
        return (d2.year - d1.year) * 12 + d2.month - d1.month

def getExcelColumnName(colId):
    '''
    colId可以是数字（1-16384），则返回对应的字母（A-XFD）
    也可以是字母（A-XFD，可以是小写），则返回对应的数字（1-16384）
    '''
    A = ord('A') # 65
    if isinstance(colId, int):
        if colId<1 or colId>16384:
            raise Exception('colId must in the range [1, 16384]')
        factor = colId
        result = ''
        while factor>0:
            remainder = factor % 26
            factor = factor // 26
            if remainder==0:
                factor -= 1
                remainder = 26
            result = chr(remainder + A - 1) + result
        return result
    elif isinstance(colId, str):
        if len(colId)>3:
            raise Exception('colId must in the range ["A", "XFD"]')
        result = 0
        for pos, letter in enumerate(list(colId.upper())[::-1]):
            if letter<'A' or letter>'Z':
                raise Exception('colId must in the range ["A", "XFD"]')
            result += (ord(letter) - A + 1) * 26**pos
        return result
    else:
        raise Exception('colId must be int in [1, 16384], or str in ["A", "XFD"]')

def excelRangeToRC(rng, asString=True):
    '''
    converts an Excel range like 'D5' to 'R5C4' format
    rng can be 'D5' or 'D5:F10' (will return 'R5C4:R10C6')
        or '5:10' (will return 'R5:R10')
        or 'D:F' (will return 'C4:C6')
    if asString, return 'R5C4', else return (5, 4)
    '''
    rng = rng.split(':')
    if len(rng) == 1:
        col = ''
        row = ''
        for d in rng[0]:
            if d>='0' and d<='9':
                row += d
            else:
                col += d
        if col!='':
            col = str(getExcelColumnName(col))
        if asString:
            return ('R' if row!='' else '') + row + ('C' if col!='' else '') + col
        else:
            return (int(row) if row!='' else np.nan, int(col) if col!='' else np.nan)
    else:
        if asString:
            return excelRangeToRC(rng[0], asString) + ':' + excelRangeToRC(rng[1], asString)
        else:
            return (excelRangeToRC(rng[0], asString), excelRangeToRC(rng[1], asString))

def excelRcToRange(rng):
    '''
    converts an Excel range like 'R5C4' to 'D5' format
    'R5C' for row 5 and 'RC4' for column 4
    '''
    rng = rng.split(':')
    if len(rng) == 1:
        row = rng[0].split('C')[0].replace('R', '')
        col = rng[0].split('C')[1]
        if col!='':
            col = getExcelColumnName(int(col))
        return col + row
    else:
        return excelRcToRange(rng[0]) + ':' + excelRcToRange(rng[1])

# Pandas相关函数------------------------------------------------------------------------
def isListEmpty(lst):
    '''
    lst can be a list or a pd.Series
    '''
    lstOfEmpty = [0, '0', '', '-', np.nan, None]
    return len(set(lst) - set(lstOfEmpty))==0

def dfAggregate(df, group_by, agg_by, agg_fun='sum', sort=None, hide_nulls=False, filters=None):
    '''
    sort = 'ascending' | 'descending' | None
    agg_by不能用'index'
    filters: {col1: cond1, con2: cond2, ...}
        cond可以是数值、str或者list，可以用*表示匹配任意值
    '''
    if filters is not None:
        for col, cond in filters.items():
            if not isinstance(cond, list):
                cond = [cond]
            if cond!=['*']:
                df = df[df[col].isin(cond)]
    if list(df.index.names)[0] is not None: # 有命名的index
        df = df.reset_index()
    df_grouped = df.groupby(group_by).agg({agg_by: agg_fun})
    df_grouped.reset_index(inplace=True)
    if sort=='descending':
        df_grouped.sort_values(agg_by, ascending=False, inplace=True)
    elif sort=='ascending':
        df_grouped.sort_values(agg_by, ascending=True, inplace=True)
    if hide_nulls:
        df_grouped = df_grouped[(df_grouped[agg_by]!=0) & ~df_grouped[agg_by].isna()]
    return df_grouped

def filterDataFrame(df, **kwargs):
    '''
    kwargs: by1 = cond1, by2 = cond2, ...，不同条件之间为and关系
    cond可以为str，写成'value'或者'>=value'、'>value'、'<=value'、'<value'的格式
    cond也可以为list，写成['value1', 'value2', ...]的格式，并且其中可以包含上述str的任意格式，list内部的条件为and关系
    '''
    index = list(df.index.names)
    if index[0] is not None: # 有命名的index
        df2 = df.reset_index()
    else:
        df2 = df.copy(deep=True)
    for key, value in kwargs.items():
        if isinstance(value, list):
            values = value
        else:
            values = [value]
        for val in values:
            if val.startswith('>='):
                df2 = df2[df2[key]>=val[2:]]
            elif val.startswith('>'):
                df2 = df2[df2[key]>val[1:]]
            elif val.startswith('<='):
                df2 = df2[df2[key]<=val[2:]]
            elif val.startswith('<'):
                df2 = df2[df2[key]<val[1:]]
            else:
                df2 = df2[df2[key]==val]
    if index[0] is not None: # 有命名的index
        df2.set_index(index, inplace=True)
    return df2

def vlookup(df_left, df_right, to_map, left_on='index', right_on='index',
    on=None, how='inner'):
    '''
    to_map、left_on、right_on可以用index中的字段（index一定要命名）
    left_on、right_on也可以直接写'index'
        可以是list或str（表示单个列）
    可以忽略left_on和right_on直接用on，表示两边用同样的字段
    how：left保留df_left所有的数据，right保留df_right所有的数据，
        inner仅保留两边都有的数据，outer全部保留
        默认为inner模式，会自动排除匹配不上的条目
    注意：
        1. 如果inner模式返回行数减少，说明存在未能匹配的内容
        2. 如果left_on和right_on存在一对多关系，返回的DataFrame行数会增加
        3. 返回的新DataFrame和df_left的排序可能不同
    '''
    left_index = list(df_left.index.names)
    right_index = list(df_right.index.names)
    if left_index[0] is not None: # 有命名的index
        df_left = df_left.reset_index()
    if df_right.index.names[0] is not None: # 有命名的index
        df_right = df_right.reset_index()
    if on is not None:
        left_on = right_on = on
    if left_on=='index':
        left_on = left_index
    if right_on=='index':
        right_on = right_index
    if not isinstance(left_on, list):
        left_on = [left_on]
    if not isinstance(right_on, list):
        right_on = [right_on]
    if not isinstance(to_map, list):
        to_map = [to_map]
    # 检查是否存在无法匹配的内容
    df_left_grouped = df_left.groupby(left_on).agg({left_on[0]: 'count'})
    df_right_grouped = df_right.groupby(right_on).agg({right_on[0]: 'count'})
    if how=='inner':
        print('check for 0: ', len(df_left_grouped[~df_left_grouped.index.isin(df_right_grouped.index)]))
    print('rows before merging:', df_left.shape[0])
    # 进行匹配
    df_left = df_left.merge(df_right[right_on + to_map],
        how=how,
        left_on=left_on,
        right_on=right_on)
    if left_index[0] is not None: # 有命名的index
        df_left = df_left.set_index(left_index)
    for col in right_on:
        if col not in left_on:
            df_left.drop(col, axis=1, inplace=True)
    # 检查匹配前后的行数变化
    print('rows after merging:', df_left.shape[0])
    return df_left

def removeDuplicates(df, by='index', keep=None):
    '''
    by: string or list, use 'index' for index, use ['index', 'col1', 'col2'...] to include index and other columns
    keep: string or list, if None then keep all columns other than by columns
    '''
    index = list(df.index.names)
    if index[0] is not None: # 有命名的index
        df2 = df.reset_index()
    else:
        df2 = df.copy(deep=True)
    if isinstance(by, list):
        if 'index' in by:
            by_cols = index + by.remove('index')
        else:
            by_cols = by
    else:
        if by=='index':
            by_cols = index
        else:
            by_cols = [by]
    if keep is None:
        keep_cols = list(df2.columns)
        for col in by_cols:
            if col in keep_cols:
                keep_cols.remove(col)
    elif isinstance(keep, list):
        keep_cols = keep
    else:
        keep_cols = [keep]
    to_agg = {}
    for col in keep_cols:
        to_agg[col] = 'first'
    df2 = df2.groupby(by_cols).agg(to_agg)
    df2.reset_index(inplace=True)
    df2_index = index.copy()
    for col in index:
        if col not in by_cols + keep_cols:
            df2_index.remove(col)
    if len(df2_index) > 0:
        df2.set_index(df2_index, inplace=True)
    else:
        df2.set_index(by, inplace=True)
    return df2

def countDistinct(lst):
    '''
    lst can be a list, Series or DataFrame
    '''
    if isinstance(lst, list):
        return len(set(lst))
    if isinstance(lst, pd.Series) or isinstance(lst, pd.DataFrame):
        return len(lst.drop_duplicates())

def dfPivot(df, rows, columns, values):
    '''
    类似Excel的pivot table，从长表转成宽表（行变成列）
    '''
    if not isinstance(rows, list):
        rows = [rows]
    if not isinstance(columns, list):
        columns = [columns]
    if not isinstance(values, list):
        values = [values]
    index = list(df.index.names)
    if index[0] is not None: # 有命名的index
        df2 = df.reset_index()
    else:
        df2 = df.copy(deep=True)
    df2 = df2[rows + columns + values]
    df2.set_index(rows + columns, inplace=True)
    df2 = df2.unstack(level=columns)
    return df2

def dfUnpivot(df, id_vars=None, value_vars=None, var_name='variable', value_name='value'):
    '''
    * melt方式：
    类似Excel的unpivot，从宽表转成长表（列变成行）
    id_vars为要保留的列，None表示除了index外都不保留
    value_vars为需要转为行的列，如果columns是MultiIndex，可以只用第一层的名称，或者用None包括所有层
    var_name为value_vars标题转为行后的列名，如果columns是MultiIndex，需要用list指定每一层转换后的列名
    value_name为value_vars内容转为行后的列名
    * stack方式的区别：
    df.stack(level=-1)要求columns为MultiIndex
    level=-1表示将最下面一级header转置成为行
    level=0表示将第一行header转置成为行
    '''
    index = list(df.index.names)
    if index[0] is not None: # 有命名的index
        df2 = df.reset_index()
    else:
        df2 = df.copy(deep=True)
    if id_vars is None and index[0] is not None:
        id_vars = index
    df2 = df2.melt(id_vars=id_vars,
        value_vars = value_vars,
        var_name = var_name,
        value_name = value_name)
    return df2

def dfFlattenMultiHeader(df, concat_with='_'):
    if not isinstance(df.columns, pd.MultiIndex):
        return df
    else:
        df2 = df.copy(deep=True)
        df2.columns = [concat_with.join(col) for col in df2.columns]
        return df2

def dfSummarize(df):
    summary = pd.DataFrame(columns=['random_sample', 'dtype', 'type',
        'non_null_cnt', 'non_null_unique_cnt',
        'min', 'max', 'mean', 'median', 'sum'])
    for col in df.columns:
        summary = summary.append(pd.Series(name=col))
        sample = df[~df[col].isna()][col].sample(n=1).iloc[0]
        summary.loc[col, 'random_sample'] = sample
        summary.loc[col, 'dtype'] = str(type(sample))[8:-2]
        summary.loc[col, 'type'] = 'text' if isinstance(sample, str) \
            else 'date' if isinstance(sample, datetime.date) \
            else 'number'
        summary.loc[col, 'non_null_cnt'] = len(df[~df[col].isna()])
        summary.loc[col, 'non_null_unique_cnt'] = df[~df[col].isna()][col].unique().size
        summary.loc[col, 'min'] = df[~df[col].isna()][col].min()
        summary.loc[col, 'max'] = df[~df[col].isna()][col].max()
        summary.loc[col, 'mean'] = df[~df[col].isna()][col].mean() if summary.loc[col, 'type']=='number' else np.nan
        summary.loc[col, 'median'] = df[~df[col].isna()][col].median() if summary.loc[col, 'type']=='number' else np.nan
        summary.loc[col, 'sum'] = df[~df[col].isna()][col].sum() if summary.loc[col, 'type']=='number' else np.nan
    return summary

def dfPrettify(df):
    TextLimit = 10
    pretty = pd.DataFrame(index=df.index, columns=df.columns)
#     if pretty.columns.contains('dtype') and pretty.columns.contains('type'):
    for rowIndex, (rowId, row) in enumerate(df.iterrows()):
        for colIndex, (colId, cell) in enumerate(row.iteritems()):
            if pd.isna(cell):
                pretty.iloc[rowIndex, colIndex] = ''
            elif isinstance(cell, str):
                if len(cell) < TextLimit:
                    pretty.iloc[rowIndex, colIndex] = cell
                else:
                    if colId=='dtype':
                        pretty.iloc[rowIndex, colIndex] = '...' + cell[-TextLimit:]
                    else:
                        pretty.iloc[rowIndex, colIndex] = cell[:TextLimit] + '...'
            elif isinstance(cell, datetime.date):
                pretty.iloc[rowIndex, colIndex] = cell
            else:
                if 0 < abs(cell) < 1:
                    pretty.iloc[rowIndex, colIndex] = '{0:,.2f}'.format(cell)
                else:
                    pretty.iloc[rowIndex, colIndex] = '{0:,.0f}'.format(cell)
    return pretty

# 接口相关函数------------------------------------------------------------------------
def getUrlData(url, token, params):
    '''
    params = {
        'created': '2018-08-13',
        'date': '20180201',
        'page': 1,
        'page_size': 10,
        'limit': 100,
        'offset': 100,
        # 其他参数例如：
        'product__name': '头条',
        'type__name': 'video',
    }
    '''
    headers = {'Authorization': 'Token {0}'.format(token)}
    try:
        data = requests.get(url, params=params, headers=headers, verify=True).json()
        return data
    except Exception as e:
        return e

def getUrlDataAll(url, token, date=None, created=None, limit=5000, page_size=5000):
    '''
    cdn、idc不需要设置date参数，默认返回所有created日期
    cloud service需要通过date取每天的数据，date默认为当天
    date/created可以设置为某一天，例如'20180319'，也可以设置为一个区间，例如['20180101', '20180830']
    date/created如果设置为['20180101']，则默认第二个日期为今天
    date和created其中之一必须为None，否则忽略date
    '''
    offset = 0
    params = {
        'page_size': page_size,
        'limit': limit,
        'offset': offset
    }
    if date is None and created is None:
        data = getUrlData(url, token, params)
        df = pd.DataFrame(data['results'])
        cnt = data['count']
        while cnt > limit:
            offset += limit
            cnt -= limit
            params['offset'] = offset
            data = getUrlData(url, token, params)
            df = pd.concat([df, pd.DataFrame(data['results'])], sort=False)
        return df
    else:
        if created is not None:
            datestr = 'created'
            datefmt = '%Y-%m-%d'
            date = created
        else:
            datestr = 'date'
            datefmt = '%Y%m%d'
        if not isinstance(date, list):
            date = [date, date]
        if len(date)==1:
            date.append(datetime.date.today())
        dates = pd.date_range(start=date[0], end=date[1], freq='D')
        df = None
        for date in dates:
            params[datestr] = date.strftime(datefmt)
            data = getUrlData(url, token, params)
            df2 = pd.DataFrame(data['results'])
            df2[datestr] = date
            print(date.date(), df2.shape)
            if df is None:
                df = df2
            else:
                df = pd.concat([df, df2], sort=False)
            cnt = data['count']
            while cnt > limit:
                offset += limit
                cnt -= limit
                params['offset'] = offset
                data = getUrlData(url, token, params)
                df2 = pd.DataFrame(data['results'])
                df2[datestr] = date
                print(date.date(), df2.shape)
                df = pd.concat([df, df2], sort=False)
        if df is None:
            return pd.DataFrame()
        else:
            return df

# 其他通用函数------------------------------------------------------------------------
def isFileExist(file):
    '''
    file can be a file or a directory
    '''
    return os.path.exists(file)

def deleteFile(file):
    return os.remove(file)

def getDirFiles(path, pattern, recursive=True):
    '''
    pattern: .xlsx, .html
    '''
    files = [os.path.join(path, i) for i in os.listdir(path) if os.path.isfile(os.path.join(path, i)) and os.path.splitext(i)[1]==pattern]
    dirs = [os.path.join(path, i) for i in os.listdir(path) if os.path.isdir(os.path.join(path, i))]
    if recursive & len(dirs)>0:
        for d in dirs:
            files.extend(getDirFiles(d, pattern, True))
    return files

def jsonPrettify(obj, indent=2, sort_keys=False):
    print(json.dumps(obj, indent=indent, sort_keys=sort_keys, ensure_ascii=False))

def formatNumberCn(n):
    if abs(n)<1: # 一以下，显示为百分比
        if round(n, 2)==round(n, 4):
            formatted = '{0:,.0%}'.format(n)
        else:
            formatted = '{0:,.2%}'.format(n)
    elif abs(n)<1e3: # 千以下
        formatted = '{0:,.2f}'.format(n)
    elif abs(n)<1e4: # 万以下
        formatted = '{0:,.0f}'.format(n)
    elif abs(n)<1e6: # 百万以下
        formatted = '{0:,.2f}万'.format(n/1e4)
    elif abs(n)<1e8: # 亿以下
        formatted = '{0:,.0f}万'.format(n/1e4)
    elif abs(n)<1e9: # 十亿以下
        formatted = '{0:,.2f}亿'.format(n/1e8)
    elif abs(n)<1e11: # 千亿以下
        formatted = '{0:,.0f}亿'.format(n/1e8)
    elif abs(n)<1e12: # 万亿以下
        formatted = '{0:,.0f}亿'.format(n/1e8)
    elif abs(n)<1e13: # 十万亿以下
        formatted = '{0:,.2f}万亿'.format(n/1e12)
    else: # 十万亿以上
        formatted = '{0:,.0f}万亿'.format(n/1e12)
    return formatted

def formatNumberEn(n):
    if abs(n)<1: # 一以下，显示为百分比
        if round(n, 2)==round(n, 4):
            formatted = '{0:,.0%}'.format(n)
        else:
            formatted = '{0:,.2%}'.format(n)
    elif abs(n)<1e3: # 千以下
        formatted = '{0:,.2f}'.format(n)
    elif abs(n)<1e4: # 万以下
        formatted = '{0:,.2f}K'.format(n/1e3)
    elif abs(n)<1e6: # 百万以下
        formatted = '{0:,.0f}K'.format(n/1e3)
    elif abs(n)<1e8: # 亿以下
        formatted = '{0:,.2f}M'.format(n/1e6)
    elif abs(n)<1e9: # 十亿以下
        formatted = '{0:,.0f}M'.format(n/1e6)
    elif abs(n)<1e11: # 千亿以下
        formatted = '{0:,.2f}B'.format(n/1e9)
    elif abs(n)<1e12: # 万亿以下
        formatted = '{0:,.0f}B'.format(n/1e9)
    elif abs(n)<1e13: # 十万亿以下
        formatted = '{0:,.2f}T'.format(n/1e12)
    else: # 十万亿以上
        formatted = '{0:,.0f}T'.format(n/1e12)
    return formatted

def getTDates(pivot=None, offset=-1, full_period=False):
    '''
    return a dict as {period_name: [period_from, period_to, days in period]}
    pivot: default None for today
    offset: default -1 for t-1
    full_period:
        False to return period_to as pivot or comparative pivot
        True to return period_to as the maximum date in that period
            (days in period will be full days too)
    本期指标：                           对应上期同比指标：
        本日：curr_day                      前日：prev_day（本日向前推7天，即上周同日）
        本周/近一周：curr_7days             前一周：prev_7days
        近一月：curr_30days                 前一月：prev_30days
        本月：curr_month（到本日）           上月：prev_month（到本日同一日期）
        本双月：curr_dmonth（到本日）        上双月：prev_dmonth（到本日同一日期）
        本年：curr_year（到本日）            上年：prev_year（到本日同一日期）
    下期指标：与上期类似，prev改为next
    返回一个字典：{指标: [起始日期, 截止日期, 天数]}
    '''
    if pivot is None:
        # datetime.datetime.today()返回的日期带有时间
        # datetime.date.today()返回的日期为date格式而不是datetime格式
        pivot = pd.to_datetime(datetime.date.today())
    pivot -= pd.tseries.offsets.DateOffset(n=-offset)
    tdates = dotdict(dict(
        curr_day = [pivot,
            pivot],
        curr_7days = [pivot - pd.tseries.offsets.DateOffset(n=6),
            pivot],
        curr_30days = [pivot - pd.tseries.offsets.DateOffset(n=29),
            pivot],
        curr_month = [bomonth(pivot),
            eomonth(pivot) if full_period else pivot],
        curr_dmonth = [bomonth(pivot, pivot.month % 2 - 1),
            eomonth(pivot, pivot.month % 2) if full_period else pivot],
        curr_year = [datetime.datetime(pivot.year, 1, 1),
            datetime.datetime(pivot.year, 12, 31) if full_period else pivot],
        prev_day = [pivot - pd.tseries.offsets.DateOffset(n=7),
            pivot - pd.tseries.offsets.DateOffset(n=7)],
        prev_7days = [pivot - pd.tseries.offsets.DateOffset(n=13),
            pivot - pd.tseries.offsets.DateOffset(n=7)],
        prev_30days = [pivot - pd.tseries.offsets.DateOffset(n=59),
            pivot - pd.tseries.offsets.DateOffset(n=30)],
        prev_month = [bomonth(pivot, -1),
            eomonth(pivot, -1) if full_period else edate(pivot, -1)],
        prev_dmonth = [bomonth(pivot, pivot.month % 2 - 3),
            eomonth(pivot, pivot.month % 2 - 2) if full_period else edate(pivot, -2)],
        prev_year = [datetime.datetime(pivot.year - 1, 1, 1),
            datetime.datetime(pivot.year - 1, 12, 31) if full_period else edate(pivot, -12)],
        next_day = [pivot + pd.tseries.offsets.DateOffset(n=7),
           pivot + pd.tseries.offsets.DateOffset(n=7)],
        next_7days = [pivot + pd.tseries.offsets.DateOffset(n=7),
            pivot + pd.tseries.offsets.DateOffset(n=13)],
        next_30days = [pivot + pd.tseries.offsets.DateOffset(n=1),
            pivot + pd.tseries.offsets.DateOffset(n=30)],
        next_month = [bomonth(pivot, 1),
            eomonth(pivot, 1) if full_period else edate(pivot, 1)],
        next_dmonth = [bomonth(pivot, pivot.month % 2 + 1),
            eomonth(pivot, pivot.month % 2 + 3) if full_period else edate(pivot, 2)],
        next_year = [datetime.datetime(pivot.year + 1, 1, 1),
            datetime.datetime(pivot.year + 1, 12, 31) if full_period else edate(pivot, 12)],
        ))
    for tdate in tdates.items():
        tdates[tdate[0]].append((tdate[1][1] - tdate[1][0]).days + 1)
    return tdates

def getDateProgress(pivot=None, period='month', as_days=False):
    '''
    return the percentage of days elapsed in certain period
    pivot: default None for today
    period: month, dmonth, year
    as_days: True to return days elapsed, false to return percentage of days elapsed
    '''
    if pivot is None:
        pivot = datetime.datetime.today()
    if period=='month':
        date_from = bomonth(pivot)
        date_to = eomonth(pivot)
    elif period=='dmonth':
        date_from = bomonth(pivot, pivot.month % 2 - 1)
        date_to = eomonth(pivot, pivot.month % 2)
    elif period=='year':
        date_from = datetime.datetime(pivot.year, 1, 1)
        date_to = datetime.datetime(pivot.year, 12, 31)
    else:
        return None
    if as_days:
        return (pivot - date_from).days + 1
    else:
        return ((pivot - date_from).days + 1) / ((date_to - date_from).days + 1)

# 自定义类------------------------------------------------------------------------
class dotdict(dict):
    '''
    dot.notation access to dictionary attributes
    '''
    __getattr__ = dict.get
    __setattr__ = dict.__setitem__
    __delattr__ = dict.__delitem__

###################################################################################################
# fpa及financedb写数相关函数------------------------------------------------------------------------
def putLog(log_dic, section='log_fin'):
    '''
    log_dic = {'section': section,
        'source': db.tbl,
        'old_cnt': 0,
        'new_cnt': 0,
        'old_sum': 0,
        'new_sum': 0,
        'total_check': 0,
        'min_date': Timestamp('2017-06-01'),
        'max_date': Timestamp('2018-07-16')}
    '''
    log = pd.Series(log_dic, name=datetime.datetime.now())
    log = pd.DataFrame(log).T
    log.loc[:, 'old_sum':'total_check'] = pd.DataFrame(log.loc[:, 'old_sum':'total_check'],
        dtype='float')
    log['min_date'] = pd.to_datetime(log['min_date'])
    log['max_date'] = pd.to_datetime(log['max_date'])
    config = get_config()
    db = config.loc[section, 'fdb']
    tbl = config.loc[section, 'ftbl']
    putToSqlTable(db, tbl, log, index=False, if_exists='append')
    print(log.T)
    return True

def putSection(section, df, append=True):
    '''
    obsolete, use putToSection instead
    df必须是按照index关键字段汇总过的
    如果append=False会先清空原表
    '''
    config = get_config()
    db = config.loc[section, 'fdb']
    tbl = config.loc[section, 'ftbl']
    index_col = config.loc[section, 'index_col']
    if index_col is not None:
        index_col = [d.strip() for d in index_col.split(sep=',')]
    df_old = getMySqlTable(db, tbl, index_col=index_col)
    if not append:
        clearMySqlTable(db, tbl)
        df_new = df
    else:
        df_new = df[~df.index.isin(df_old.index)]
    if len(df_new) > 0:
        putToSqlTable(db, tbl, df_new,
            index=True if index_col is not None else False, if_exists='append')
    # generate log
    log = {'section': section,
        'source': db + '.' + tbl,
        'old_cnt': len(df_old),
        'new_cnt': len(df)}
    if 'value' in df.columns:
        value_col = 'value'
    elif 'launch_user' in df.columns:
        value_col = 'launch_user'
    else:
        value_col = None
    if value_col is not None:
        log['old_sum'] = round(df_old[value_col].sum(),2) if len(df_old) > 0 else 0
        log['new_sum'] = round(df[value_col].sum(),2)
        log['total_check'] = round((log['new_sum'] - log['old_sum'] - df_new[value_col].sum()),2) \
            if append else round((log['new_sum'] - log['old_sum']),2)
    else:
        log['old_sum'] = np.nan
        log['new_sum'] = np.nan
        log['total_check'] = np.nan
    if 'date' in df.reset_index().columns:
        date_col = 'date'
    elif 'p_date' in df.reset_index().columns:
        date_col = 'p_date'
    else:
        date_col = None
    if date_col is not None:
        log['min_date'] = df.reset_index()[date_col].min()
        log['max_date'] = df.reset_index()[date_col].max()
    else:
        log['min_date'] = np.nan
        log['max_date'] = np.nan
    putLog(log)
    print(log)
    return df_new.shape

def getSection(section, fdb=True, to_clipboard=False):
    '''
    obsolte, use getGoogleSection instead
    '''
    config = get_config()
    db = config.loc[section, 'fdb' if fdb else 'db']
    tbl = config.loc[section, 'ftbl' if fdb else 'tbl']
    index_col = config.loc[section, 'index_col']
    index_col = [d.strip() for d in index_col.split(sep=',')]
    index_col = index_col if fdb else None
    df = getMySqlTable(db, tbl, index_col=index_col)
    if to_clipboard:
        df.to_clipboard(sep=',', index=True)
    return df

def putToSection(section, df, append=True, put_log=True):
    '''
    df必须是按照index_col关键字段汇总过的，并且set_index(index_col)过
    如果append=False会先清空原表
    如果append=True会根据index_col去重后写入
    如果没有index_col或者index_col为默认的id列，则无法自动去重，需要在写入之前手工去重
    '''
    config = get_config()
    db = config.loc[section, 'fdb']
    tbl = config.loc[section, 'ftbl']
    index_col = config.loc[section, 'index_col']
    if index_col=='id' and df.index.name!='id':
        index_col = None
    if index_col is not None:
        index_col = [d.strip() for d in index_col.split(sep=',')]
    df_old = getToSqlTable(db, tbl, index_col=index_col)
    if not append:
        clearToSqlTable(db, tbl)
        df_new = df
    else: # append
        # 如果没有index_col或者index_col为默认的id列，则无法自动去重，需要在写入之前手工去重
        if index_col is not None:
            df_new = df[~df.index.isin(df_old.index)]
        else:
            df_new = df
    if len(df_new) > 0:
        putToSqlTable(db, tbl, df_new,
            index=True if index_col is not None else False, if_exists='append')
    # generate log
    if put_log:
        log = {'section': section,
            'source': db + '.' + tbl,
            'old_cnt': len(df_old),
            'new_cnt': len(df)} # 没有用df_new，显示的是传过来的数，而不是写进去的数
        if 'value' in df.columns:
            value_col = 'value'
        elif 'launch_user' in df.columns:
            value_col = 'launch_user'
        else:
            value_col = None
        if value_col is not None:
            log['old_sum'] = round(df_old[value_col].sum(),2) if len(df_old) > 0 else 0
            log['new_sum'] = round(df[value_col].sum(),2)
            log['total_check'] = round((log['new_sum'] - log['old_sum'] - df_new[value_col].sum()),2) \
                if append else round((log['new_sum'] - log['old_sum']),2)
        else:
            log['old_sum'] = np.nan
            log['new_sum'] = np.nan
            log['total_check'] = np.nan
        date_col = None
        for d in ['date', 'p_date', 'datetime', 'created', 'updated']:
            if d in df.reset_index().columns:
                date_col = d
        if date_col is not None:
            log['min_date'] = df.reset_index()[date_col].min()
            log['max_date'] = df.reset_index()[date_col].max()
        else:
            log['min_date'] = np.nan
            log['max_date'] = np.nan
        putLog(log)
    return df_new.shape

def getMapping(config_name, to_map, df, left_on='index', right_on='index'):
    cfg = getGoogleSection(config_name)
    print(config_name, cfg.shape)
    left_key = df.index if left_on=='index' else df[left_on]
    right_key = cfg.index if right_on=='index' else cfg[right_on]
    if isinstance(to_map, list):
        right_map = cfg[to_map] if right_on=='index' else cfg[[right_on] + to_map]
    else:
        right_map = cfg[[to_map]] if right_on=='index' else cfg[[right_on, to_map]]
    # mapping check
    print('unmapped count:', df[~left_key.isin(right_key)][left_on].unique().size)
    print('unmapped items:', df[~left_key.isin(right_key)][left_on].unique())
    # mapping
    return df.merge(right=right_map,
        how='left',
        left_index = True if left_on=='index' else False,
        right_index = True if right_on=='index' else False,
        left_on=None if left_on=='index' else left_on,
        right_on=None if right_on=='index' else right_on)

# MySQL相关函数（fpa）------------------------------------------------------------------------
def getMySqlConnection(db):
    dbs = get_dbs()
    conn = pymysql.connect(host=dbs.loc[db].host,
        port=int(dbs.loc[db].port),
        db=dbs.loc[db]['name'],
        user=dbs.loc[db].user,
        passwd=dbs.loc[db].passwd,
        autocommit=True,
        charset='utf8')
    return conn
def getMySqlTable(db, tbl, index_col=None, cols=None, cond=None):
    '''
    cols: None表示*，即所有列，如果需要选择列，传入[col1, col2, ...]
    cond: 不需要写where，直接写例如：month='2018-09-01' or month='2018-10-01'
    '''
    conn = getMySqlConnection(db)
    if cols is None:
        cols = '*'
    else:
        cols = ', '.join(cols)
    sql = 'select ' + cols + 'from ' + tbl
    if cond is not None:
        sql += ' where ' + cond
    df = pd.read_sql(sql, conn, index_col=index_col)
    conn.close()
    return df
def useMySqlDb(db):
    conn = getMySqlConnection(db)
    cursor = conn.cursor()
    sql = 'use ' + db
    cursor.execute(sql)
    conn.commit()
    return conn, cursor
def dropMySqlTable(db, tbl):
    conn, cursor = useMySqlDb(db)
    sql = 'drop table if exists ' + tbl
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    return True
def clearMySqlTable(db, tbl, cond=None):
    conn, cursor = useMySqlDb(db)
    sql = 'delete from ' + tbl
    if cond is not None:
        sql += ' where ' + cond
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    return True
def createMySqlTableFrom(db, tbl, new_tbl, cols=[]):
    conn, cursor = dropMySqlTable(db, new_tbl)
    if len(cols)==0:
        cols_to_select = '*'
    else:
        cols_to_select = ', '.join(cols)
    sql = 'create table ' + new_tbl + \
        ' select ' + cols_to_select + ' from ' + tbl
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    return True

# MySQL相关函数（financedb）------------------------------------------------------------------------
def getToSqlConnection(db):
    dbs = get_dbs()
    yconnect_str = 'mysql+mysqldb://' + dbs.loc[db].user + ':' + dbs.loc[db].passwd + \
        '@' + dbs.loc[db].host + ':' + str(dbs.loc[db].port) + \
        '/' + dbs.loc[db]['name'] + '?charset=utf8&autocommit=true'
    yconnect = create_engine(yconnect_str)
    return yconnect
def getToSqlTable(db, tbl, index_col=None, cols=None, cond=None):
    '''
    cols: None表示*，即所有列，如果需要选择列，传入[col1, col2, ...]
    cond: 不需要写where，直接写例如：month='2018-09-01' or month='2018-10-01'
    '''
    yconnect = getToSqlConnection(db)
    if cols is None:
        cols = '*'
    else:
        cols = ', '.join(cols)
    sql = 'select ' + cols + ' from ' + tbl
    if cond is not None:
        sql += ' where ' + cond
    df = pd.read_sql(sql, yconnect, index_col=index_col)
    yconnect.dispose()
    return df
def putToSqlTable(db, tbl, df, index=False, if_exists='append'):
    '''
    if_exists: fail, replace, append
    '''
    yconnect = getToSqlConnection(db)
    df.to_sql(tbl, yconnect, index=index, if_exists=if_exists)
    yconnect.dispose()
    return True
def clearToSqlTable(db, tbl, cond=None):
    '''
    cond: 不需要写where，直接写例如：month='2018-09-01' or month='2018-10-01'
    '''
    yconnect = getToSqlConnection(db)
    sql = 'delete from ' + tbl
    if cond is not None:
        sql += ' where ' + cond
    try:
        pd.read_sql_query(sql, yconnect)
        return True
    except:
        return False
    finally:
        yconnect.dispose()

# Sqlite相关函数------------------------------------------------------------------------
def getSqliteConnection(db='data/init_fbi.db'):
    conn = sqlite3.connect(db)
    return conn
def getSqliteTable(tbl, db='data/init_fbi.db', index_col=None, cols=None, cond=None):
    '''
    cols: None表示*，即所有列，如果需要选择列，传入[col1, col2, ...]
    cond: 不需要写where，直接写例如：month='2018-09-01' or month='2018-10-01'
    '''
    conn = getSqliteConnection(db)
    if cols is None:
        cols = '*'
    else:
        cols = ', '.join(cols)
    sql = 'select ' + cols + ' from ' + tbl
    if cond is not None:
        sql += ' where ' + cond
    df = pd.read_sql(sql, conn, index_col=index_col)
    conn.commit()
    conn.close()
    return df
def putSqliteTable(tbl, df, db='data/init_fbi.db', index=False, if_exists='append'):
    '''
    if_exists: fail, replace, append
    '''
    conn = getSqliteConnection(db)
    df.to_sql(tbl, conn, index=index, if_exists=if_exists)
    conn.commit()
    conn.close()
    return True
def clearSqliteTable(tbl, db='data/init_fbi.db', cond=None):
    '''
    cond: 不需要写where，直接写例如：month='2018-09-01' or month='2018-10-01'
    '''
    conn = getSqliteConnection(db)
    sql = 'delete from ' + tbl
    if cond is not None:
        sql += ' where ' + cond
    try:
        pd.read_sql_query(sql, conn)
        conn.commit()
        return True
    except:
        return False
    finally:
        conn.close()

# Google Sheets setup------------------------------------------------------------------------
def setupService():
    '''
    先运行data目录下的googlesheets.py产生credentials.json文件
    in case of failed connection, try up to 10 times
    '''
    from googleapiclient import discovery
    from httplib2 import Http
    from oauth2client import file, client, tools

    SCOPES = 'https://www.googleapis.com/auth/spreadsheets' #.readonly
    store = file.Storage('data/credentials.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('data/client_secret.json', SCOPES)
        creds = tools.run_flow(flow, store)
    for i in range(10):
        try:
            service = discovery.build('sheets', 'v4', http=creds.authorize(Http()))
            break
        except Exception as e:
            print(e)
            print('trying ' + str(i + 2))
    return service

def getGoogleSheetId(url):
    return url.replace('https://docs.google.com/spreadsheets/d/', '').split('/')[0]

def getGoogleSheet(sheetId, tabName, address=None, index_col=None,
    valueRenderOption='UNFORMATTED_VALUE',
    returnDf=True, service=None):
    '''
    address: 'A1:D10' format or 'R1C1:R10C4' format, None for all used range
    valueRenderOption: FORMATTED_VALUE, UNFORMATTED_VALUE, FORMULA
    in case of failed connection, try up to 10 times
    right-most column can't be empty with only a title, or without any titile
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    if address is not None:
        address = "'" + tabName + "'!" + address
    else:
        address = tabName
    for i in range(10):
        try:
            result = service.spreadsheets().values().get(spreadsheetId=sheetId,
                range=address,
                valueRenderOption=valueRenderOption).execute()
            break
        except Exception as e:
            print(e)
            print('trying ' + str(i + 2))
    values = result.get('values', [])
    if returnDf:
        if not values:
            return pd.DataFrame()
        df = pd.DataFrame(data=values[1:], columns=values[0])
        df.dropna(axis=0, how='all', inplace=True)
        if index_col is not None:
            df.set_index(index_col, inplace=True)
        return df
    else:
        return values # 右边的空列和下面的空行默认不返回

def putGoogleSheetCell(sheetId, tabName, row, column, value,
    valueInputOption='USER_ENTERED',
    service=None):
    '''
    same as putGoogleSheetRange when values is a single value,
        but use row and column index instead of a range name
    valueInputOption: RAW, USER_ENTERED
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    address = "'" + tabName + "'!R" + str(row) + 'C' + str(column)
    return service.spreadsheets().values().update(spreadsheetId=sheetId,
        range=address,
        body={'values': [[value]]},
        valueInputOption=valueInputOption).execute()

def putGoogleSheetRange(sheetId, tabName, address, values,
    valueInputOption='USER_ENTERED',
    service=None):
    '''
    address: 'A1:D10' format or 'R1C1:R10C4' format
    values: [[r1c1, r1c2, ...], ...]，也可以是唯一值，一个list，或者一个DataFrame或Series
    如果values是唯一值：
        日期时间格式会被自动转换为'yyyy-mm-dd'形式
        np.nan会清除当前单元格内容
    如果values是嵌套列表，其中的日期时间数值需要自行转换
    如果values是DataFrame，会忽略index和column，转换成嵌套列表后写入
        转换时np.nan和np.inf会被替换成空字符串
    valueInputOption: RAW, USER_ENTERED
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'updatedRange': "'20181204'!E6:R59",
     'updatedRows': 54,
     'updatedColumns': 14,
     'updatedCells': 756}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    if isinstance(values, datetime.datetime): # values是唯一值并且为日期时间格式
        values = '{0:%Y-%m-%d}'.format(values)
    if isinstance(values, pd.Series):
        values = values.replace(np.nan, '')
        values = values.replace(np.inf, '')
        values = values.replace(-np.inf, '')
        if values.dtype=='<M8[ns]':
            values = values.astype(str)
        values = [[d for d in values]]
    elif isinstance(values, pd.DataFrame):
        values = values.replace(np.nan, '')
        values = values.replace(np.inf, '')
        values = values.replace(-np.inf, '')
        for col in values.columns:
            if values[col].dtype=='<M8[ns]':
                values[col] = values[col].astype(str)
        values = [[d for d in row] for _, row in values.iterrows()]
    elif isinstance(values, list):
        if not isinstance(values[0], list): # 非嵌套列表
            values = [values]
    else: # 写入唯一值
        values = [[values]]
    address = "'" + tabName + "'!" + address
    return service.spreadsheets().values().update(spreadsheetId=sheetId,
        range=address,
        body={'values': values},
        valueInputOption=valueInputOption).execute()

def putDfToGoogleSheet(sheetId, tabName, df,
    setTabRed=False, setFilter=True, setFrozen=True, service=None):
    '''
    sheetId必须是已经存在的Google Sheet的id，可以是完整链接
    如果tabName不存在，新建tabName表
        如果已经存在则删除多余行和多余列
    df目前只能是一行表头
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    sheetId = sheetId.replace('https://docs.google.com/spreadsheets/d/', '')
    sheetId = sheetId.split('/')[0]
    if isGoogleSheetTabExist(sheetId, tabName, service=service):
        print(tabName, 'exists')
        tab_size = getGoogleSheetTabSize(sheetId, tabName, service=service)
        if tab_size[0] > df.shape[0] + 1: # 表头占一行
            address = 'R' + str(df.shape[0] + 1) + 'C:R' + str(tab_size[0]) + 'C'
            # Goolge Sheet开了筛选后无法删除部分行
            clearGoogleSheetTabFilter(sheetId, tabName, service=service)
            deleteGoogleSheetRange(sheetId, tabName, address=address, service=service)
            print(address + ' deleted') # 提示的address其实为0-based
        if tab_size[1] > df.shape[1]:
            address = 'RC' + str(df.shape[1]) + ':RC' + str(tab_size[1]) # [左包含，右不不包含)
            # Goolge Sheet开了筛选后无法删除部分行
            clearGoogleSheetTabFilter(sheetId, tabName, service=service)
            deleteGoogleSheetRange(sheetId, tabName, address=address, service=service)
            print(address + ' deleted') # 提示的address其实为0-based
    else:
        addGoogleSheetTab(sheetId, tabName, service=service)
        print(tabName, 'added')
    if setTabRed:
        setGoogleSheetTabColor(sheetId, tabName, service=service)
    # 写标题行
    values = ['{0:%Y-%m-%d}'.format(value)
        if isinstance(value, datetime.datetime) else value
        for value in df.columns]
    print(putGoogleSheetRange(sheetId, tabName, 'A1', values, service=service))
    # 写数据
    print(putGoogleSheetRange(sheetId, tabName, 'A2', df, service=service))
    if setFilter:
        setGoogleSheetTabFilter(sheetId, tabName, service=service)
    if setFrozen:
        freezeGoogleSheetTab(sheetId, tabName, service=service)

def getGoogleSheetWorkbook(sheetId, service=None):
    '''
    return the workbook as a dictionary
    in case of failed connection, try up to 10 times
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    for i in range(10):
        try:
            workbook = service.spreadsheets().get(spreadsheetId=sheetId).execute()
            break
        except Exception as e:
            print(e)
            print('trying ' + str(i + 2))
    return workbook

def getGoogleSheetTabs(sheetId, service=None):
    '''
    return the titles of the sheets as a list
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    workbook = getGoogleSheetWorkbook(sheetId, service=service)
    tabs = [tab['properties']['title'] for tab in workbook.get('sheets', [])]
    print(len(tabs))
    return tabs

def getGoogleSheetName(sheetId, service=None):
    '''
    return the title of the spreadsheet
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    workbook = getGoogleSheetWorkbook(sheetId, service=service)
    name = list(workbook.values())[1]['title']
    return name

def renameGoogleSheet(sheetId, newName, service=None):
    '''
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'updateSpreadsheetProperties': {
            'fields': 'title',
            'properties': {'title': newName}
        }}]}).execute()

def getGoogleSheetTabId(sheetId, tabName, service=None):
    '''
    return the sheetId of the sheet (tab)
    sheetId is the the id of each sheet in a spreadsheet, i.e. 1742198400
    spreadsheetId is the id of the spreadsheet, i.e. 1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    workbook = getGoogleSheetWorkbook(sheetId, service=service)
    for sheet in workbook['sheets']:
        if sheet['properties']['title'] == tabName:
            return sheet['properties']['sheetId']
    return None

def getGoogleSheetTabSize(sheetId, tabName, dataOnly=True, service=None):
    '''
    return the size of the sheet (tab) as (rows, columns)
    if not dataOnly, return the full size including empty rows and columns
    sheetId is the the id of each sheet in a spreadsheet, i.e. 1742198400
    spreadsheetId is the id of the spreadsheet, i.e. 1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    workbook = getGoogleSheetWorkbook(sheetId, service=service)
    for sheet in workbook['sheets']:
        if sheet['properties']['title'] == tabName:
            if not dataOnly:
                return (sheet['properties']['gridProperties']['rowCount'],
                    sheet['properties']['gridProperties']['columnCount'])
            else:
                data = getGoogleSheet(sheetId, tabName, returnDf=False, service=service)
                return(len(data), max([len(r) for r in data]))
    return False

def isGoogleSheetTabHidden(sheetId, tabName, service=None):
    '''
    return the hidden status of the sheet (tab)
    sheetId is the the id of each sheet in a spreadsheet, i.e. 1742198400
    spreadsheetId is the id of the spreadsheet, i.e. 1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    workbook = getGoogleSheetWorkbook(sheetId, service=service)
    for sheet in workbook['sheets']:
        if sheet['properties']['title'] == tabName:
            return sheet['properties'].get('hidden', False)
    return None

def isGoogleSheetTabExist(sheetId, tabName, service=None):
    '''
    return the hidden status of the sheet (tab)
    sheetId is the the id of each sheet in a spreadsheet, i.e. 1742198400
    spreadsheetId is the id of the spreadsheet, i.e. 1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    workbook = getGoogleSheetWorkbook(sheetId, service=service)
    for sheet in workbook['sheets']:
        if sheet['properties']['title'] == tabName:
            return True
    return False

def copyGoogleSheetTab(sheetId, tabName, copyToSheetId=None, service=None):
    '''
    if copyToSheetId is None, duplicate the tab in the same spreadsheet
    return copied sheet as:
    {'sheetId': 546728693,
     'title': 'template（副本）',
     'index': 2,
     'sheetType': 'GRID',
     'gridProperties': {'rowCount': 60,
        'columnCount': 18,
        'frozenRowCount': 5,
        'frozenColumnCount': 4,
        'hideGridlines': True,
        'rowGroupControlAfter': True,
        'columnGroupControlAfter': True}}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    if copyToSheetId is None:
        copyToSheetId = sheetId
    else:
        copyToSheetId = getGoogleSheetId(copyToSheetId)
    return service.spreadsheets().sheets().copyTo(spreadsheetId=sheetId,
        sheetId=shtId,
        body={'destinationSpreadsheetId': copyToSheetId}).execute()

def duplicateGoogleSheetTab(sheetId, tabName, newTabName, index=0, service=None):
    '''
    same as copyGoogleSheetTab when copyToSheetId is None,
        with the capability of rename the new tab
        and insert in a designated position
    index: The zero-based index where the new sheet should be inserted.
        The index of all sheets after this are incremented.
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{'duplicateSheet': {'properties': {'sheetId': 1584923354,
        'title': 'template2',
        'index': 0,
        'sheetType': 'GRID',
        'gridProperties': {'rowCount': 60,
            'columnCount': 18,
            'frozenRowCount': 5,
            'frozenColumnCount': 4,
            'hideGridlines': True,
            'rowGroupControlAfter': True,
            'columnGroupControlAfter': True}}}}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'duplicateSheet': {
            'sourceSheetId': shtId,
            'newSheetName': newTabName,
            'insertSheetIndex': index
        }}]}).execute()

def addGoogleSheetTab(sheetId, tabName, index=0, service=None):
    '''
    add a new tab and insert in a designated position
    index: The zero-based index where the new sheet should be inserted.
        The index of all sheets after this are incremented.
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{'duplicateSheet': {'properties': {'sheetId': 1584923354,
        'title': 'template2',
        'index': 0,
        'sheetType': 'GRID',
        'gridProperties': {'rowCount': 60,
            'columnCount': 18,
            'frozenRowCount': 5,
            'frozenColumnCount': 4,
            'hideGridlines': True,
            'rowGroupControlAfter': True,
            'columnGroupControlAfter': True}}}}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'addSheet': {
            'properties': {'index': index,
                'title': tabName}
        }}]}).execute()

def setGoogleSheetTabColor(sheetId, tabName, tabColor={'red': 1,
    'green': 0,
    'blue': 0,
    'alpha': 1}, service=None):
    '''
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'updateSheetProperties': {
            'fields': 'tabColor',
            'properties': {'sheetId': shtId,
                'tabColor': {'red': tabColor['red'],
                    'green': tabColor['green'],
                    'blue': tabColor['blue'],
                    'alpha': tabColor['alpha']}}
        }}]}).execute()

def renameGoogleSheetTab(sheetId, tabName, newTabName, service=None):
    '''
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'updateSheetProperties': {
            'fields': 'title',
            'properties': {'sheetId': shtId,
                'title': newTabName}
        }}]}).execute()

def reindexGoogleSheetTab(sheetId, tabName, index=0, service=None):
    '''
    move the tab to a new position
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'updateSheetProperties': {
            'fields': 'index',
            'properties': {'sheetId': shtId,
                'index': index}
        }}]}).execute()

def deleteGoogleSheetTab(sheetId, tabName, service=None):
    '''
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'deleteSheet': {
            'sheetId': shtId
        }}]}).execute()

def deleteGoogleSheetRange(sheetId, tabName, address='',
    shiftDimension='ROWS', service=None):
    '''
    Deletes a range of cells, shifting other cells into the deleted area.
    address: A1, A1:B1, A:B, 1:2, empty ('') to delete everything on a tab
        R1C1 format is acceptable too
    shiftDimension: ROWS to shift upward, COLUMNS to shift left
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    tab_size = getGoogleSheetTabSize(sheetId, tabName, service=service)
    if address=='':
        startRowIndex = 0
        endRowIndex = tab_size[0] # end index is exclusive -- [start_index, end_index)
        startColumnIndex = 0
        endColumnIndex = tab_size[1] # end index is exclusive -- [start_index, end_index)
    else:
        if address.find(':')<0:
            address = address + ':' + address
        if address.split(':')[0].startswith('R') and 'C' in address.split(':')[0][1:]: # R1C1 format
            address = excelRcToRange(address)
        rc_address = excelRangeToRC(address, asString=False)
        startRowIndex = 0 if np.isnan(rc_address[0][0]) else rc_address[0][0]
        endRowIndex = tab_size[0] if np.isnan(rc_address[1][0]) else rc_address[1][0]
        startColumnIndex = 0 if np.isnan(rc_address[0][1]) else rc_address[0][1]
        endColumnIndex = tab_size[1] if np.isnan(rc_address[1][1]) else rc_address[1][1]
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'deleteRange': {
            'range': {'sheetId': shtId,
                'startRowIndex': startRowIndex,
                'endRowIndex': endRowIndex,
                'startColumnIndex': startColumnIndex,
                'endColumnIndex': endColumnIndex},
            'shiftDimension': shiftDimension
        }}]}).execute()

def resizeGoogleSheetRange(sheetId, tabName, address, keepRows, service=None):
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    old_data = getGoogleSheet(sheetId, tabName, address)
    if old_data.shape[0] > keepRows - 1:
        old_size = excelRangeToRC(address, asString=False)
        address_to_delete = 'R' + str(keepRows + 1) + \
            'C' + str(old_size[0][1]) + \
            ':R' + str(old_data.shape[0] + 1) + \
            'C' + str(old_size[1][1])
        cells = []
        for row in range(keepRows + 1, old_data.shape[0] + 2):
            cells.append([''] * (old_size[1][1] - old_size[0][1] + 1))
        putGoogleSheetRange(sheetId, tabName, address_to_delete, cells, service=service)
#         deleteGoogleSheetRange(sheetId, tabName, address_to_delete, service=service)
        print(address_to_delete, 'deleted')

def hideGoogleSheetTab(sheetId, tabName, hidden=True, service=None):
    '''
    use hidden=False to unhide a tab
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'updateSheetProperties': {
            'fields': 'hidden',
            'properties': {'sheetId': shtId,
                'hidden': hidden}
        }}]}).execute()

def setGoogleSheetTabFilter(sheetId, tabName, service=None):
    '''
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    tab_size = getGoogleSheetTabSize(sheetId, tabName, service=service)
    startRowIndex = 0
    endRowIndex = tab_size[0] # end index is exclusive -- [start_index, end_index)
    startColumnIndex = 0
    endColumnIndex = tab_size[1] # end index is exclusive -- [start_index, end_index)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'setBasicFilter': {
            'filter': {'range': {'sheetId': shtId,
                'startRowIndex': startRowIndex,
                'endRowIndex': endRowIndex,
                'startColumnIndex': startColumnIndex,
                'endColumnIndex': endColumnIndex}}
        }}]}).execute()

def clearGoogleSheetTabFilter(sheetId, tabName, service=None):
    '''
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'clearBasicFilter': {
            'sheetId': shtId
        }}]}).execute()

def freezeGoogleSheetTab(sheetId, tabName,
    frozenRowCount=1, frozenColumnCount=1, service=None):
    '''
    use frozenRowCount=0, frozenColumnCount=0 to unfreeze
    return a dictionary as:
    {'spreadsheetId': '1D5oNDGgCWjX1yoO6ai2IlxtDaGEkMnAfwOK1O3BOWVo',
     'replies': [{}]}
    '''
    sheetId = getGoogleSheetId(sheetId)
    if service is None:
        service = setupService()
    shtId = getGoogleSheetTabId(sheetId, tabName, service=service)
    tab_size = getGoogleSheetTabSize(sheetId, tabName, service=service)
    return service.spreadsheets().batchUpdate(spreadsheetId=sheetId,
        body={'requests': [{'updateSheetProperties': {
            'fields': 'gridProperties',
            'properties': {'sheetId': shtId,
                'gridProperties': {
                    'rowCount': tab_size[0],
                    'columnCount': tab_size[1],
                    'frozenRowCount': frozenRowCount,
                    'frozenColumnCount': frozenColumnCount}
                    }
        }}]}).execute()

# fpa相关------------------------------------------------------------------------
def getGoogleSection(section, fdb=True, to_clipboard=False):
    config = get_config()
    db = config.loc[section, 'fdb' if fdb else 'db']
    tbl = config.loc[section, 'ftbl' if fdb else 'tbl']
    index_col = config.loc[section, 'index_col']
    if index_col is not None:
        index_col = [d.strip() for d in index_col.split(sep=',')]
    index_col = index_col if fdb else None
    if db=='ggl_cfg': # get from Google Config
        df = getGoogleSheet(FDB_CONFIG, tbl, index_col=index_col)
    elif db=='ggl_ipt': # get from Google Input
        df = getGoogleSheet(FDB_INPUT, tbl, index_col=index_col)
    else: # get from MySQL
        df = getMySqlTable(db, tbl, index_col=index_col)
    df.replace('FALSE', False, inplace=True)
    df.replace('TRUE', True, inplace=True)
    df.replace('', np.nan, inplace=True)
    if to_clipboard:
        df.to_clipboard(sep=',', index=True)
    return df

def getGoogleMapping(config_name, to_map, df, left_on='index', right_on='index'):
    cfg = getGoogleSection(config_name)
    print(config_name, cfg.shape)
    if left_on=='index':
        left_key = df.index
    else:
        left_key = df[left_on]
    if right_on=='index':
        cfg.index = cfg.index.astype(left_key.dtype)
        right_key = cfg.index
    else:
        cfg[right_on] = cfg[right_on].astype(left_key.dtype)
        right_key = cfg[right_on]
    if isinstance(to_map, list):
        right_map = cfg[to_map] if right_on=='index' else cfg[[right_on] + to_map]
    else:
        right_map = cfg[[to_map]] if right_on=='index' else cfg[[right_on, to_map]]
    # mapping check
    print('unmapped count:', df[~left_key.isin(right_key)][left_on].unique().size)
    print('unmapped items:', df[~left_key.isin(right_key)][left_on].unique())
    # mapping
    return df.merge(right=right_map,
        how='left',
        left_index = True if left_on=='index' else False,
        right_index = True if right_on=='index' else False,
        left_on=None if left_on=='index' else left_on,
        right_on=None if right_on=='index' else right_on)

def pivot_report(df, rows, cols):
    '''
    cols = [curr_week_days, curr_week_sum,
        prev_week_days, prev_week_sum,
        mtd_days, mtd_sum,
        ytd_days, ytd_sum]
    curr_week_days = dotdict(dict(name = '本周天数',
        date_from = t1.curr_week_from,
        date_to = t1.curr_week_to,
        date_col = 'date',
        value_col = 'date',
        agg = countDistinct))
    '''
    df2 = df[rows].drop_duplicates().set_index(rows)
    for col in cols:
        df_filtered = df[(df[col.date_col] >= col.date_from) &
            (df[col.date_col] <= col.date_to)]
        if col.value_col!='date':
            df_filtered = df_filtered.groupby([col.date_col] + rows).agg({col.value_col: 'sum'})
            df_filtered.reset_index(inplace=True)
        df_grouped = df_filtered.groupby(rows).agg({col.value_col: col.agg})
        df2[col.name] = df_grouped[col.value_col]
    return df2

def get_dbs():
    return getExcel(DBINFO, 'dbs')

def get_config():
    index_col = 'section'
    return getGoogleSheet(FDB_CONFIG, FDB_CONFIG_TAB, index_col=index_col)
    # print('\t' + getGoogleSheetName(FDB_CONFIG) + ' ' + FDB_CONFIG_TAB + ': ' + str(config.shape))

# Lark相关------------------------------------------------------------------------
def sendLarkMessage(msg, title=None, emails=None, chat_ids=None):
    '''
    emails or chat_ids should at least provide one
    can be a list or string (for one single email or chat_id)
    '''
    from lark import LarkBot
    from async_lark import AsyncLarkBot
    if not isinstance(emails, list):
        emails = [emails]
    if not isinstance(chat_ids, list):
        chat_ids = [chat_ids]
    bot = LarkBot(bot_token='b-cb831711-4173-47de-9178-db9ae1231ab6')
    '''
    user_id = bot.email2user_id("liaoshian")['user_id']
    bot.user_info(user_id) returns
    {'code': 0,
     'msg': 'Success',
     'ok': True,
     'id': '6545303538883363076',
     'name': '廖世安',
     'en_name': 'Shian Liao',
     'profile_pic': 'https://sf1-ttcdn-tos.pstatp.com/obj/mosaic-legacy/834a0014477c519a4b7a'}
    '''
    # reformat the message
    text = msg.replace('\n', '').replace('  ', '')
    if title is None:
        msg_type = LarkBot.MsgType.TEXT
        content = {'text': text}
    else:
        msg_type = LarkBot.MsgType.POST # RICH_TEXT会丢失正文
        content = {
            'title': title,
            'text': text
        }
    if emails!=[None]:
        for email in emails:
            bot.private_chat(email=email,
                msg_type=msg_type,
                content=content)
            print('message sent to', email)
    if chat_ids!=[None]:
        for chat_id in chat_ids:
            bot.message(chat_id=chat_id,
                msg_type=msg_type,
                content=content)
            print('message sent to', chat_id)